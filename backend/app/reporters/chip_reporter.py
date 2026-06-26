"""ChipReporter — 每日籌碼分析日報 Excel 生成器.

每個交易日 17:30 由 scheduler 觸發，或 Admin API 手動觸發。
輸出到 CHIP/data/reports/YYYYMMDD_chip_report.xlsx。

Sheets
------
1. 總覽      — 日期、市場快照、外資前五大買超/賣超
2. 三大法人  — 全部股票，依合計淨買超降冪
3. 外資連買  — foreign_streak != 0，依連買天數降冪（正→負）
4. 融資融券  — 有融資券餘額的股票，依融資增減降冪
5. 期貨走勢  — 近 30 日 TXF/MXF 三大法人未平倉淨口數
"""
from __future__ import annotations

import logging
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.config import settings

logger = logging.getLogger("chip.reporter")

# ── Palette ──────────────────────────────────────────────────────────────────
_NAVY   = "1E3A5F"
_WHITE  = "FFFFFF"
_POS_BG = "DBEAFE"   # light blue  — 買超/正數
_NEG_BG = "FEE2E2"   # light red   — 賣超/負數
_STR_POS = "DCFCE7"  # light green — 連買
_STR_NEG = "FFEDD5"  # light orange — 連賣
_GRAY   = "F1F5F9"
_BORDER_COLOR = "CBD5E1"

_THIN = Side(style="thin", color=_BORDER_COLOR)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def _hdr_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_NAVY)

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _hdr_font() -> Font:
    return Font(bold=True, color=_WHITE, name="Calibri", size=11)

def _num_font(bold: bool = False) -> Font:
    return Font(bold=bold, name="Calibri", size=10)

def _center(ws: Any, row: int, col: int, value: Any, **kw) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal="center", vertical="center", **kw)
    c.border = _BORDER
    return c

def _right(ws: Any, row: int, col: int, value: Any) -> Any:
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = _BORDER
    return c

def _autowidth(ws: Any, min_w: int = 8, max_w: int = 30) -> None:
    for col in ws.columns:
        width = min_w
        for cell in col:
            try:
                v = str(cell.value or "")
                # Chinese chars count as 2
                w = sum(2 if ord(c) > 127 else 1 for c in v) + 2
                width = max(width, w)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, max_w)


# ── ChipReporter ─────────────────────────────────────────────────────────────

class ChipReporter:

    def __init__(self, db: Session):
        self.db = db

    def generate(self, target_date: date) -> Path:
        """Generate Excel daily report. Returns output file path."""
        logger.info("generating chip report for %s", target_date)

        wb = Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        self._sheet_overview(wb, target_date)
        self._sheet_institutional(wb, target_date)
        self._sheet_streaks(wb, target_date)
        self._sheet_margin(wb, target_date)
        self._sheet_futures(wb, target_date)

        fname = f"{target_date.strftime('%Y%m%d')}_chip_report.xlsx"
        fpath = settings.reports_dir / fname
        wb.save(fpath)
        logger.info("report saved: %s", fpath)
        return fpath

    # ── Sheet helpers ─────────────────────────────────────────────────────────

    def _write_header_row(self, ws: Any, row: int, headers: list[str]) -> None:
        for col, hdr in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=hdr)
            c.fill = _hdr_fill()
            c.font = _hdr_font()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _BORDER

    def _color_signed(self, ws: Any, row: int, col: int, value: int | float | None) -> None:
        if value is None:
            ws.cell(row=row, column=col, value="—")
            return
        c = _right(ws, row, col, value)
        c.font = _num_font()
        c.number_format = '#,##0'
        if value > 0:
            c.fill = _fill(_POS_BG)
        elif value < 0:
            c.fill = _fill(_NEG_BG)

    def _color_streak(self, ws: Any, row: int, col: int, streak: int) -> None:
        c = _center(ws, row, col, streak)
        c.font = _num_font(bold=abs(streak) >= 3)
        if streak >= 3:
            c.fill = _fill(_STR_POS)
        elif streak <= -3:
            c.fill = _fill(_STR_NEG)

    # ── Sheet 1: 總覽 ──────────────────────────────────────────────────────

    def _sheet_overview(self, wb: Workbook, d: date) -> None:
        ws = wb.create_sheet("總覽")
        ws.sheet_view.showGridLines = False

        # Title
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = f"台灣籌碼日報 — {d.strftime('%Y/%m/%d')}"
        c.font = Font(bold=True, size=16, color=_NAVY, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30

        # ── Futures snapshot ──
        row = 3
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"].value = "▸ 期貨未平倉快照"
        ws[f"A{row}"].font = Font(bold=True, size=12, color=_NAVY)
        row += 1

        fut_rows = self.db.execute(text("""
            SELECT contract,
                   foreign_long - foreign_short AS foreign_net,
                   trust_long - trust_short AS trust_net,
                   dealer_long - dealer_short AS dealer_net,
                   oi_total
            FROM raw_futures_oi
            WHERE date = :d
            ORDER BY contract
        """), {"d": str(d)}).fetchall()

        if fut_rows:
            self._write_header_row(ws, row, ["合約", "外資淨口數", "投信淨口數", "自營淨口數", "全市場未平倉"])
            row += 1
            for r in fut_rows:
                ws.cell(row=row, column=1, value=r[0]).border = _BORDER
                for ci, v in enumerate([r[1], r[2], r[3]], 2):
                    self._color_signed(ws, row, ci, v)
                c5 = _right(ws, row, 5, r[4])
                c5.number_format = "#,##0"
                row += 1
        else:
            ws[f"A{row}"].value = "（無期貨資料）"
            row += 1

        # ── Top 5 foreign buy ──
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"].value = "▸ 外資前五大買超"
        ws[f"A{row}"].font = Font(bold=True, size=12, color=_NAVY)
        row += 1

        top_buy = self.db.execute(text("""
            SELECT p.stock_id, s.name, p.foreign_net, p.foreign_streak,
                   p.margin_balance, p.margin_ratio
            FROM processed_chip p
            LEFT JOIN stocks s ON s.stock_id = p.stock_id
            WHERE p.date = :d AND p.foreign_net > 0
            ORDER BY p.foreign_net DESC LIMIT 5
        """), {"d": str(d)}).fetchall()

        self._write_header_row(ws, row, ["代號", "名稱", "外資淨(張)", "外資連買天", "融資餘額", "資券比%"])
        row += 1
        for r in top_buy:
            ws.cell(row=row, column=1, value=r[0]).border = _BORDER
            ws.cell(row=row, column=2, value=r[1]).border = _BORDER
            self._color_signed(ws, row, 3, r[2])
            self._color_streak(ws, row, 4, r[3])
            c5 = _right(ws, row, 5, r[4])
            c5.number_format = "#,##0"
            _right(ws, row, 6, round(r[5], 2) if r[5] else None)
            row += 1

        # ── Top 5 foreign sell ──
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"].value = "▸ 外資前五大賣超"
        ws[f"A{row}"].font = Font(bold=True, size=12, color=_NAVY)
        row += 1

        top_sell = self.db.execute(text("""
            SELECT p.stock_id, s.name, p.foreign_net, p.foreign_streak,
                   p.margin_balance, p.margin_ratio
            FROM processed_chip p
            LEFT JOIN stocks s ON s.stock_id = p.stock_id
            WHERE p.date = :d AND p.foreign_net < 0
            ORDER BY p.foreign_net ASC LIMIT 5
        """), {"d": str(d)}).fetchall()

        self._write_header_row(ws, row, ["代號", "名稱", "外資淨(張)", "外資連買天", "融資餘額", "資券比%"])
        row += 1
        for r in top_sell:
            ws.cell(row=row, column=1, value=r[0]).border = _BORDER
            ws.cell(row=row, column=2, value=r[1]).border = _BORDER
            self._color_signed(ws, row, 3, r[2])
            self._color_streak(ws, row, 4, r[3])
            c5 = _right(ws, row, 5, r[4])
            c5.number_format = "#,##0"
            _right(ws, row, 6, round(r[5], 2) if r[5] else None)
            row += 1

        _autowidth(ws)
        ws.column_dimensions["B"].width = 16
        ws.freeze_panes = "A2"

    # ── Sheet 2: 三大法人 ──────────────────────────────────────────────────

    def _sheet_institutional(self, wb: Workbook, d: date) -> None:
        ws = wb.create_sheet("三大法人")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        headers = ["代號", "名稱", "外資淨(張)", "投信淨(張)", "自營淨(張)",
                   "合計淨(張)", "外資連買天", "投信連買天"]
        self._write_header_row(ws, 1, headers)

        rows = self.db.execute(text("""
            SELECT p.stock_id, s.name,
                   p.foreign_net, p.trust_net, p.dealer_net, p.total_net,
                   p.foreign_streak, p.trust_streak
            FROM processed_chip p
            LEFT JOIN stocks s ON s.stock_id = p.stock_id
            WHERE p.date = :d
            ORDER BY p.total_net DESC
        """), {"d": str(d)}).fetchall()

        for ri, r in enumerate(rows, 2):
            bg = _GRAY if ri % 2 == 0 else _WHITE
            ws.cell(ri, 1, r[0]).border = _BORDER
            c2 = ws.cell(ri, 2, r[1])
            c2.border = _BORDER
            if ri % 2 == 0:
                ws.cell(ri, 1).fill = _fill(bg)
                c2.fill = _fill(bg)
            for ci, val in enumerate([r[2], r[3], r[4], r[5]], 3):
                self._color_signed(ws, ri, ci, val)
            self._color_streak(ws, ri, 7, r[6])
            self._color_streak(ws, ri, 8, r[7])

        _autowidth(ws)
        ws.column_dimensions["B"].width = 16

    # ── Sheet 3: 外資連買 ──────────────────────────────────────────────────

    def _sheet_streaks(self, wb: Workbook, d: date) -> None:
        ws = wb.create_sheet("外資連買")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        headers = ["代號", "名稱", "外資淨(張)", "外資連買天", "投信淨(張)",
                   "投信連買天", "融資餘額", "資券比%"]
        self._write_header_row(ws, 1, headers)

        rows = self.db.execute(text("""
            SELECT p.stock_id, s.name,
                   p.foreign_net, p.foreign_streak,
                   p.trust_net, p.trust_streak,
                   p.margin_balance, p.margin_ratio
            FROM processed_chip p
            LEFT JOIN stocks s ON s.stock_id = p.stock_id
            WHERE p.date = :d AND p.foreign_streak != 0
            ORDER BY p.foreign_streak DESC
        """), {"d": str(d)}).fetchall()

        for ri, r in enumerate(rows, 2):
            ws.cell(ri, 1, r[0]).border = _BORDER
            ws.cell(ri, 2, r[1]).border = _BORDER
            self._color_signed(ws, ri, 3, r[2])
            self._color_streak(ws, ri, 4, r[3])
            self._color_signed(ws, ri, 5, r[4])
            self._color_streak(ws, ri, 6, r[5])
            c7 = _right(ws, ri, 7, r[6])
            c7.number_format = "#,##0"
            _right(ws, ri, 8, round(r[7], 2) if r[7] else None)

        _autowidth(ws)
        ws.column_dimensions["B"].width = 16

    # ── Sheet 4: 融資融券 ──────────────────────────────────────────────────

    def _sheet_margin(self, wb: Workbook, d: date) -> None:
        ws = wb.create_sheet("融資融券")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        headers = ["代號", "名稱", "融資餘額", "融資增減", "融券餘額", "融券增減", "資券比%"]
        self._write_header_row(ws, 1, headers)

        rows = self.db.execute(text("""
            SELECT p.stock_id, s.name,
                   p.margin_balance, p.margin_change,
                   p.short_balance, p.short_change,
                   p.margin_ratio
            FROM processed_chip p
            LEFT JOIN stocks s ON s.stock_id = p.stock_id
            WHERE p.date = :d AND (p.margin_balance > 0 OR p.short_balance > 0)
            ORDER BY p.margin_change DESC
        """), {"d": str(d)}).fetchall()

        for ri, r in enumerate(rows, 2):
            ws.cell(ri, 1, r[0]).border = _BORDER
            ws.cell(ri, 2, r[1]).border = _BORDER
            for ci, val in [(3, r[2]), (5, r[4])]:
                c = _right(ws, ri, ci, val)
                c.number_format = "#,##0"
                c.font = _num_font()
            self._color_signed(ws, ri, 4, r[3])
            self._color_signed(ws, ri, 6, r[5])
            _right(ws, ri, 7, round(r[6], 2) if r[6] else None)

        _autowidth(ws)
        ws.column_dimensions["B"].width = 16

    # ── Sheet 5: 期貨走勢 ──────────────────────────────────────────────────

    def _sheet_futures(self, wb: Workbook, d: date) -> None:
        ws = wb.create_sheet("期貨走勢")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        start = (d - timedelta(days=60)).isoformat()
        headers = ["日期",
                   "TXF 外資淨(口)", "TXF 投信淨(口)", "TXF 自營淨(口)",
                   "MXF 外資淨(口)", "MXF 投信淨(口)", "MXF 自營淨(口)"]
        self._write_header_row(ws, 1, headers)

        # Pivot: one row per date, TXF and MXF side by side
        rows_raw = self.db.execute(text("""
            SELECT date, contract,
                   foreign_long - foreign_short AS foreign_net,
                   trust_long - trust_short AS trust_net,
                   dealer_long - dealer_short AS dealer_net
            FROM raw_futures_oi
            WHERE date >= :s AND date <= :e AND contract IN ('TXF', 'MXF')
            ORDER BY date DESC, contract
        """), {"s": start, "e": str(d)}).fetchall()

        # Build dict: date → {TXF: {...}, MXF: {...}}
        pivot: dict[str, dict] = {}
        for r in rows_raw:
            pivot.setdefault(r[0], {})[r[1]] = {
                "foreign": r[2], "trust": r[3], "dealer": r[4]
            }

        for ri, (dt, contracts) in enumerate(sorted(pivot.items(), reverse=True), 2):
            ws.cell(ri, 1, dt).border = _BORDER
            txf = contracts.get("TXF", {})
            mxf = contracts.get("MXF", {})
            self._color_signed(ws, ri, 2, txf.get("foreign"))
            self._color_signed(ws, ri, 3, txf.get("trust"))
            self._color_signed(ws, ri, 4, txf.get("dealer"))
            self._color_signed(ws, ri, 5, mxf.get("foreign"))
            self._color_signed(ws, ri, 6, mxf.get("trust"))
            self._color_signed(ws, ri, 7, mxf.get("dealer"))

        _autowidth(ws)
